"""Ejecuta en lote los jobs pendientes de Twitter respetando una pausa."""

from __future__ import annotations

import os
import re
import sys
import subprocess
import time
from pathlib import Path
from dataclasses import dataclass
from typing import List, Optional, Set

import psycopg


@dataclass
class Settings:
    limit: int = int(os.getenv("RUN_TWITTER_LIMIT", "10"))
    sleep_seconds: float = float(os.getenv("RUN_TWITTER_SLEEP", "30"))
    dsn: str = os.getenv(
        "POSTGRES_DSN",
        "dbname=reto_scraper user=reto_writer password=Ale211083 host=localhost",
    )
    excel_path: str = os.getenv(
        "TWITTER_EXCEL_PATH",
        "/Users/alejandroyankilevich/Documents/MASTER DATA SCIENCE/Clases/RETO/Medios/Medios_Andalucia_Web_Redes_Final.xlsx",
    )


def extract_twitter_usernames_from_excel(excel_path: str) -> Set[str]:
    """
    Extrae los nombres de usuario de Twitter/X de la columna E del Excel.
    
    Args:
        excel_path: Ruta al archivo Excel
        
    Returns:
        Set de nombres de usuario (sin @ y sin URLs)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl no está instalado. Instálalo con: pip install openpyxl"
        )
    
    usernames = set()
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # La columna E es la columna 5
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=5).value
        if not cell_value or not isinstance(cell_value, str):
            continue
        
        # Extraer usuario de URLs como https://x.com/usuario o https://twitter.com/usuario
        # También puede ser solo el usuario o @usuario
        # Primero intentar extraer de URL completa
        url_pattern = r'https?://(?:x\.com|twitter\.com)/([a-zA-Z0-9_]+)'
        url_match = re.search(url_pattern, cell_value, re.IGNORECASE)
        if url_match:
            username = url_match.group(1).lower().strip()
            if username and username not in ['x', 'twitter']:
                usernames.add(username)
        else:
            # Si no es URL, buscar usuario directo (puede tener @ o no)
            # Solo tomar si parece un nombre de usuario válido (sin espacios, sin caracteres especiales)
            direct_pattern = r'@?([a-zA-Z0-9_]{1,15})\b'
            matches = re.findall(direct_pattern, cell_value, re.IGNORECASE)
            for match in matches:
                username = match.lower().strip()
                if username and username not in ['x', 'twitter'] and len(username) >= 1:
                    usernames.add(username)
    
    wb.close()
    return usernames


def fetch_pending_jobs(
    conn,
    limit: int,
    allowed_usernames: Optional[Set[str]] = None,
    require_mode: bool = True,
) -> List[tuple[str, str]]:
    """Obtiene jobs pendientes de Twitter.

    Por defecto, solo toma jobs creados para el modo ReTo (timeline_replies/handle_replies),
    evitando ejecutar jobs legacy basados en términos (que consumen cuota y no sirven).

    Args:
        conn: Conexión a la base de datos
        limit: Límite de jobs a obtener
        allowed_usernames: Handles permitidos (sin @). Si se provee, se filtra por igualdad exacta.
        require_mode: Si True, exige metadata.search.mode en {'timeline_replies','handle_replies'}.

    Returns:
        Lista de tuplas (job_id, term)
    """

    mode_clause = ""
    params: list = []
    if require_mode:
        mode_clause = "AND COALESCE(metadata->'search'->>'mode','') IN ('timeline_replies','handle_replies')"

    if allowed_usernames:
        # Normalizar a minúsculas y filtrar por igualdad exacta: term = handle
        handles = sorted({u.lower().lstrip('@').strip() for u in allowed_usernames if u})
        sql = f"""
            SELECT job_id::text, term
            FROM crawl_jobs
            WHERE network = 'twitter'
              AND status = 'pending'
              {mode_clause}
              AND lower(term) = ANY(%s)
            ORDER BY created_at
            LIMIT %s;
        """
        params = [handles, limit]
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return [(row[0], row[1]) for row in cur.fetchall()]

    # Sin filtro de handles: igual exigimos modo ReTo para no ejecutar legacy
    sql = f"""
        SELECT job_id::text, term
        FROM crawl_jobs
        WHERE network = 'twitter'
          AND status = 'pending'
          {mode_clause}
        ORDER BY created_at
        LIMIT %s;
    """
    with conn.cursor() as cur:
        cur.execute(sql, (limit,))
        return [(row[0], row[1]) for row in cur.fetchall()]


def run_job(job_id: str) -> tuple[bool, str]:
    """
    Ejecuta un job y retorna si fue exitoso y el motivo si falló.
    
    Returns:
        tuple: (success: bool, reason: str)
    """
    # Asegurar que usamos el mismo Python que está ejecutando este script
    python_exe = sys.executable
    # Cambiar al directorio del proyecto (donde está reto-scraper)
    project_root = Path(__file__).parent.parent
    cmd = [python_exe, "-m", "jobs.run_job", "--job-id", job_id]
    # Ejecutar desde el directorio raíz del proyecto para que encuentre el módulo jobs
    result = subprocess.run(cmd, cwd=project_root, check=False, capture_output=True, text=True)
    
    if result.returncode != 0:
        # Verificar si es un error de cuota
        error_output = result.stderr + result.stdout
        if "quota exceeded" in error_output.lower() or "UsageCapExceeded" in error_output:
            return False, "quota_exceeded"
        return False, "error"
    return True, "success"


def main() -> None:
    settings = Settings()
    
    # Leer usuarios permitidos del Excel
    print(f"Leyendo perfiles de Twitter del archivo Excel: {settings.excel_path}")
    try:
        allowed_usernames = extract_twitter_usernames_from_excel(settings.excel_path)
        print(f"Se encontraron {len(allowed_usernames)} perfiles únicos en el Excel.")
        if len(allowed_usernames) > 0:
            print(f"Primeros 10 perfiles: {list(allowed_usernames)[:10]}")
    except Exception as e:
        print(f"⚠️ Error al leer el Excel: {e}")
        print("Continuando sin filtro de usuarios...")
        allowed_usernames = None
    
    print(f"\nConectando a PostgreSQL con DSN: {settings.dsn!r}")
    with psycopg.connect(settings.dsn) as conn:
        jobs = fetch_pending_jobs(conn, settings.limit, allowed_usernames, require_mode=True)

    if not jobs:
        if allowed_usernames:
            print("No se encontraron jobs pendientes de Twitter en modo ReTo (timeline_replies) que coincidan con los perfiles del Excel.")
        else:
            print("No se encontraron jobs pendientes de Twitter en modo ReTo (timeline_replies).")
        return

    print(f"\nSe ejecutarán {len(jobs)} jobs (pausa {settings.sleep_seconds}s entre cada uno).")
    quota_exceeded = False
    
    for job_id, term in jobs:
        if quota_exceeded:
            print(f"\n⚠️ Cuota excedida. Deteniendo ejecución.")
            current_idx = next(i for i, (j, _) in enumerate(jobs) if j == job_id)
            print(f"   Jobs restantes: {len(jobs) - current_idx - 1}/{len(jobs)}")
            break
            
        print(f"\n=== Ejecutando job {job_id} (término: {term}) ===")
        success, reason = run_job(job_id)
        
        if not success and reason == "quota_exceeded":
            print(f"\n❌ ERROR: Cuota mensual de Twitter API excedida.")
            print(f"   Deteniendo procesamiento. Espera hasta el próximo ciclo mensual.")
            quota_exceeded = True
            break
        
        time.sleep(settings.sleep_seconds)
    
    if quota_exceeded:
        print(f"\n⚠️ RESUMEN: Se detuvo por cuota excedida.")
        print(f"   Revisa tu plan de Twitter API y espera al próximo ciclo mensual.")


if __name__ == "__main__":
    main()
